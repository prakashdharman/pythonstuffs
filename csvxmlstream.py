import xlrd
import socket

# TCP listener configuration
TCP_HOST = '127.0.0.1'
TCP_PORT = 8888

def stream_data_to_tcp(data, tcp_host, tcp_port):
    try:
        # Create a TCP socket
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            # Connect to the TCP server
            s.connect((tcp_host, tcp_port))
            
            # Send data
            s.sendall(data.encode())
            
            # Receive acknowledgment if needed
            # ack = s.recv(1024)
            # print("Received:", ack.decode())
            
            # Close the connection
            s.close()
    except Exception as e:
        print("Error:", e)

def main():
    xls_file = 'data.xls'  # Path to the Excel file
    sheet_name = 'Sheet1'   # Name of the sheet
    column_index = 2        # Index of the column (0-indexed)
    
    try:
        # Open the XLS file
        workbook = xlrd.open_workbook(xls_file)
        sheet = workbook.sheet_by_name(sheet_name)
        
        # Iterate through each row and extract data from the specified column
        for row_idx in range(1, sheet.nrows):  # Skip header row
            data = sheet.cell_value(row_idx, column_index)
            
            # Stream data to TCP listener
            stream_data_to_tcp(data, TCP_HOST, TCP_PORT)
            
            # Optional: You can add a delay between sending each data
            # time.sleep(1)
    except FileNotFoundError:
        print(f"File '{xls_file}' not found.")
    except Exception as e:
        print("Error:", e)




import openpyxl
import socket

# TCP listener configuration
TCP_HOST = '127.0.0.1'
TCP_PORT = 8888

def stream_data_to_tcp(data, tcp_host, tcp_port):
    try:
        # Create a TCP socket
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            # Connect to the TCP server
            s.connect((tcp_host, tcp_port))
            
            # Send data
            s.sendall(data.encode())
            
            # Receive acknowledgment if needed
            # ack = s.recv(1024)
            # print("Received:", ack.decode())
            
            # Close the connection
            s.close()
    except Exception as e:
        print("Error:", e)

def main():
    xlsx_file = 'data.xlsx'  # Path to the Excel file
    sheet_name = 'Sheet1'    # Name of the sheet
    column_index = 2         # Index of the column (0-indexed)
    
    try:
        # Open the XLSX file
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook[sheet_name]
        
        # Iterate through each row and extract data from the specified column
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            data = row[column_index]
            
            # Stream data to TCP listener
            stream_data_to_tcp(str(data), TCP_HOST, TCP_PORT)
            
            # Optional: You can add a delay between sending each data
            # time.sleep(1)
    except FileNotFoundError:
        print(f"File '{xlsx_file}' not found.")
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    main()

if __name__ == "__main__":
    main()




import openpyxl
import socket

# TCP listener configuration
TCP_HOST = '127.0.0.1'
TCP_PORT = 8888

def stream_data_to_tcp(data, tcp_host, tcp_port):
    try:
        # Create a TCP socket
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            # Connect to the TCP server
            s.connect((tcp_host, tcp_port))
            
            # Send data
            s.sendall(data.encode())
            
            # Receive acknowledgment if needed
            # ack = s.recv(1024)
            # print("Received:", ack.decode())
            
            # Close the connection
            s.close()
    except Exception as e:
        print("Error:", e)

def main():
    xlsx_file = 'data.xlsx'  # Path to the Excel file
    sheet_name = 'Sheet1'    # Name of the sheet
    column_index = 2         # Index of the column (0-indexed)
    
    try:
        # Open the XLSX file
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook[sheet_name]
        
        # Check if the column index is valid
        if column_index >= len(sheet[1]):
            print("Error: Column index is out of range")
            return
        
        # Iterate through each row and extract data from the specified column
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            data = row[column_index]
            
            # Stream data to TCP listener
            stream_data_to_tcp(str(data), TCP_HOST, TCP_PORT)
            
            # Optional: You can add a delay between sending each data
            # time.sleep(1)
    except FileNotFoundError:
        print(f"File '{xlsx_file}' not found.")
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    main()


import openpyxl
import socket

# TCP listener configuration
TCP_HOST = '127.0.0.1'
TCP_PORT = 8888

def stream_data_to_tcp(data, tcp_host, tcp_port):
    try:
        # Create a TCP socket
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            # Connect to the TCP server
            s.connect((tcp_host, tcp_port))
            
            # Send data
            s.sendall(data.encode())
            
            # Receive acknowledgment if needed
            # ack = s.recv(1024)
            # print("Received:", ack.decode())
            
            # Close the connection
            s.close()
    except Exception as e:
        print("Error:", e)

def main():
    xlsx_file = 'data.xlsx'  # Path to the Excel file
    sheet_name = 'Sheet1'    # Name of the sheet
    column_index = 2         # Index of the column (0-indexed)
    
    try:
        # Open the XLSX file
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook[sheet_name]
        
        # Check the number of columns in the sheet
        num_columns = sheet.max_column
        
        # Check if the column index is valid
        if column_index >= num_columns:
            print("Error: Column index is out of range")
            return
        
        # Iterate through each row and extract data from the specified column
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            data = row[column_index]
            
            # Stream data to TCP listener
            stream_data_to_tcp(str(data), TCP_HOST, TCP_PORT)
            
            # Optional: You can add a delay between sending each data
            # time.sleep(1)
    except FileNotFoundError:
        print(f"File '{xlsx_file}' not found.")
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    main()








import openpyxl
import socket

# TCP listener configuration
TCP_HOST = '127.0.0.1'
TCP_PORT = 8888

def stream_data_to_tcp(data, tcp_host, tcp_port):
    try:
        # Create a TCP socket
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            # Connect to the TCP server
            s.connect((tcp_host, tcp_port))
            
            # Send data
            s.sendall(data.encode())
            
            # Receive acknowledgment if needed
            # ack = s.recv(1024)
            # print("Received:", ack.decode())
            
            # Close the connection
            s.close()
    except Exception as e:
        print("Error:", e)

def main():
    xlsx_file = 'data.xlsx'  # Path to the Excel file
    sheet_name = 'Sheet1'    # Name of the sheet
    column_letter = 'C'      # Letter of the column (e.g., 'A', 'B', 'C', etc.)
    
    try:
        # Open the XLSX file
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook[sheet_name]
        
        # Iterate through each row and extract data from the specified column
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            data = row[sheet[f"{column_letter}1"].column - 1]
            
            # Stream data to TCP listener
            stream_data_to_tcp(str(data), TCP_HOST, TCP_PORT)
            
            # Optional: You can add a delay between sending each data
            # time.sleep(1)
    except FileNotFoundError:
        print(f"File '{xlsx_file}' not found.")
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    main()





import openpyxl
import socket
import ssl

# TCP server configuration
TCP_HOST = '127.0.0.1'
TCP_PORT = 8888

def stream_data_to_tcp(data, tcp_host, tcp_port):
    try:
        # Create a TCP socket
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        
        # Wrap the socket with SSL
        ssl_socket = ssl.wrap_socket(s, ssl_version=ssl.PROTOCOL_TLS)
        
        # Connect to the TCP server
        ssl_socket.connect((tcp_host, tcp_port))
        
        # Send data
        ssl_socket.sendall(data.encode())
        
        # Close the connection
        ssl_socket.close()
    except Exception as e:
        print("Error:", e)

def main():
    xlsx_file = 'data.xlsx'  # Path to the Excel file
    sheet_name = 'Sheet1'    # Name of the sheet
    column_letter = 'C'      # Letter of the column (e.g., 'A', 'B', 'C', etc.)
    
    try:
        # Open the XLSX file
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook[sheet_name]
        
        # Iterate through each row and extract data from the specified column
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            data = row[sheet[f"{column_letter}1"].column - 1]
            
            # Stream data over TLS to TCP server
            stream_data_to_tcp(str(data), TCP_HOST, TCP_PORT)
            
            # Optional: You can add a delay between sending each data
            # time.sleep(1)
    except FileNotFoundError:
        print(f"File '{xlsx_file}' not found.")
    except Exception as e:
        print("Error:", e)
if __name__ == "__main__":
    main()





# all at once

import openpyxl
import socket
import ssl

# TCP server configuration
TCP_HOST = '127.0.0.1'
TCP_PORT = 8888

def stream_data_to_tcp(data, tcp_socket):
    try:
        # Send data in chunks with chunked transfer encoding
        chunk_size = 1024
        for i in range(0, len(data), chunk_size):
            chunk = data[i:i+chunk_size]
            response = f"{len(chunk):X}\r\n{chunk}\r\n"
            tcp_socket.sendall(response.encode())
        
        # Signal end of transmission
        tcp_socket.sendall(b"0\r\n\r\n")
    except Exception as e:
        print("Error:", e)

def main():
    xlsx_file = 'data.xlsx'  # Path to the Excel file
    sheet_name = 'Sheet1'    # Name of the sheet
    column_letter = 'C'      # Letter of the column (e.g., 'A', 'B', 'C', etc.)
    
    try:
        # Open the XLSX file
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook[sheet_name]
        
        # Prepare all data to be streamed
        all_data = ""
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            data = row[sheet[f"{column_letter}1"].column - 1]
            all_data += str(data) + "\n"  # Add newline delimiter between events
        
        # Create a TCP socket
        server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        
        # Bind the socket to the address and port
        server_socket.bind((TCP_HOST, TCP_PORT))
        
        # Listen for incoming connections
        server_socket.listen(5)
        print(f"Listening on {TCP_HOST}:{TCP_PORT}...")
        
        while True:
            # Accept incoming connection
            client_socket, addr = server_socket.accept()
            print(f"Accepted connection from {addr[0]}:{addr[1]}")
            
            # Wrap the socket with SSL
            ssl_socket = ssl.wrap_socket(client_socket, server_side=True, certfile="server.crt", keyfile="server.key", ssl_version=ssl.PROTOCOL_TLS)
            
            # Stream all data to TCP client
            stream_data_to_tcp(all_data, ssl_socket)
            
            # Close the SSL socket
            ssl_socket.close()
        
        # Close the server socket
        server_socket.close()
    
    except FileNotFoundError:
        print(f"File '{xlsx_file}' not found.")
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    main()
